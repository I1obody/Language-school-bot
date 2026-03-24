import asyncio
import os
import re
import threading
import time
import zipfile
from datetime import datetime, timedelta
import pytz
import json

import pandas as pd
from aiogram import Bot, Dispatcher, F, Router
from aiogram.enums import ParseMode
from aiogram.filters import Command
from aiogram.types import (CallbackQuery, InlineKeyboardButton,
                           InlineKeyboardMarkup, Message, ReplyKeyboardMarkup,
                           KeyboardButton, FSInputFile, ReplyKeyboardRemove)
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import ValidationError
from aiogram.exceptions import TelegramBadRequest

import config

# КОНСТАНТЫ И ВРЕМЕННОЕ ХРАНИЛИЩЕ
DB_FILE = 'db/clients.xlsx'
user_data = {}
admin_requests_cache = {}

LANGS = [{"Английский": 0}, {"Китайский": 1}, {"Немецкий": 2}, {"Французский": 3}, {"Испанский": 4}, {"Японский": 5}, {"Итальянский": 6}]
DELIVERY_METHODS = [{"Очно": 0}, {"Онлайн": 1}]
GOALS = [{"Для школы": 0}, {"Для работы": 1}, {"Нравится язык": 2}, {"Переезд в другую страну": 3}, {"Для путешествий": 4}, {"Поддержание уровня": 5}, {"Другое": 6}]
GROUPS = [{"Индивидуальный": 0}, {"Группа": 2}]
AGES = [{"Малыши (3-4 года)": 0}, {"Дети (5-6 лет)": 1}, {"Школьники (7-17 лет)": 2}, {"Взрослые (18+ лет)": 3}]
LEVELS = [{"С нуля": 0}, {"Начальный (A1)": 1}, {"Элементарный (A2)": 2}, {"Средний (B1)": 3}, {"Промежуточный (B2)": 4}, {"Выше среднего (C1)": 5}, {"Продвинутый (C2)": 6}]

REVERSE_LANGS = {str(v): k for d in LANGS for k, v in d.items()}
REVERSE_DELIVERY_METHODS = {str(v): k for d in DELIVERY_METHODS for k, v in d.items()}
REVERSE_GOALS = {str(v): k for d in GOALS for k, v in d.items()}
REVERSE_GROUPS = {str(v): k for d in GROUPS for k, v in d.items()}
REVERSE_AGES = {str(v): k for d in AGES for k, v in d.items()}
REVERSE_LEVELS = {str(v): k for d in LEVELS for k, v in d.items()}


# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# Добавляем временную зону
MOSCOW_TZ = pytz.timezone(config.ADMIN_WORKING_HOURS['timezone'])

def is_within_working_hours():
    """
    Проверка на отработку в часы работы персонала.
    Понедельник-пятница: 10:00 - 22:00
    Суббота-воскресенье: 10:00 - 16:00
    """
    now = datetime.now(MOSCOW_TZ)
    weekday = now.weekday()
    hour = now.hour

    if 0 <= weekday <= 4:  # Monday to Friday
        return config.ADMIN_WORKING_HOURS['mon_fri_start'] <= hour < config.ADMIN_WORKING_HOURS['mon_fri_end']
    elif 5 <= weekday <= 6:  # Saturday and Sunday
        return config.ADMIN_WORKING_HOURS['sat_sun_start'] <= hour < config.ADMIN_WORKING_HOURS['sat_sun_end']
    
    return False

def add_to_queue(request_data):
    """Добавляет новый запрос в файл очереди."""
    if not os.path.exists('db'):
        os.makedirs('db')
    
    queue = []
    if os.path.exists(config.REQUEST_QUEUE_FILE):
        with open(config.REQUEST_QUEUE_FILE, 'r', encoding='utf-8') as f:
            queue = json.load(f)

    queue.append(request_data)

    with open(config.REQUEST_QUEUE_FILE, 'w', encoding='utf-8') as f:
        json.dump(queue, f, ensure_ascii=False, indent=4)

def clear_queue():
    """Очищает файл очереди запросов."""
    if os.path.exists(config.REQUEST_QUEUE_FILE):
        os.remove(config.REQUEST_QUEUE_FILE)

def escape_markdown_v2(text: str) -> str:
    """Экранирует все специальные символы в строке для MarkdownV2."""
    escape_chars = r'\_*[]()~`>#+-=|{}.!'
    return re.sub(f'([{re.escape(escape_chars)}])', r'\\\1', str(text))


def decode_selection(callback_data: str) -> str:
    """Расшифровывает строку callback_data в читаемый текст."""
    parts = re.findall(r'&([a-z]+)(\d+)', callback_data)
    decoded_parts = []
    for marker, value in sorted(parts):
        if marker == 'l':
            decoded_parts.append(REVERSE_LANGS.get(value, ''))
        elif marker == 'd':
            decoded_parts.append(REVERSE_DELIVERY_METHODS.get(value, ''))
        elif marker == 'o':
            decoded_parts.append(REVERSE_GOALS.get(value, ''))
        elif marker == 'g':
            decoded_parts.append(REVERSE_GROUPS.get(value, ''))
        elif marker == 'a':
            decoded_parts.append(REVERSE_AGES.get(value, ''))
        elif marker == 'e':
            decoded_parts.append(REVERSE_LEVELS.get(value, ''))
    return ", ".join(filter(None, decoded_parts))


async def _send_inbox_messages(chat_id: int, bot: Bot, index: int = 0, message_id: int = None):
    """
    Отправляет менеджеру одну необработанную заявку с кнопками навигации.
    Использует message_id для редактирования существующего сообщения.
    """
    cache_key = 'unprocessed_requests'
    
    # Проверяем, есть ли данные в кэше. Если нет, загружаем из файла.
    if cache_key not in admin_requests_cache:
        try:
            df = pd.read_excel(DB_FILE)
            unprocessed_requests = df[df['Статус'] == 'Необработана']
            admin_requests_cache[cache_key] = unprocessed_requests
        except FileNotFoundError:
            try:
                await bot.send_message(chat_id=chat_id, text="Файл с заявками не найден.")
            except (TelegramBadRequest, ValidationError) as e:
                print(f"Ошибка при отправке сообщения 'Файл с заявками не найден': {e}")
            return

    requests_df = admin_requests_cache[cache_key]
    
    if requests_df.empty:
        text = "📥 Все заявки обработаны\\. Список пуст\\."
        if message_id:
            try:
                await bot.edit_message_text(chat_id=chat_id, message_id=message_id, text=text, parse_mode=ParseMode.MARKDOWN_V2)
            except (TelegramBadRequest, ValidationError) as e:
                print(f"Ошибка при редактировании сообщения 'Список пуст': {e}")
                try:
                    await bot.send_message(chat_id=chat_id, text=f"Произошла ошибка при обновлении\\. Подробности смотрите в консоли\\.", parse_mode=ParseMode.MARKDOWN_V2)
                except (TelegramBadRequest, ValidationError) as e2:
                    print(f"Ошибка при отправке сообщения об ошибке: {e2}")
        else:
            try:
                await bot.send_message(chat_id=chat_id, text=text, parse_mode=ParseMode.MARKDOWN_V2)
            except (TelegramBadRequest, ValidationError) as e:
                print(f"Ошибка при отправке сообщения 'Список пуст': {e}")
        return

    total_requests = len(requests_df)
    
    if not 0 <= index < total_requests:
        index = 0

    request = requests_df.iloc[index]
    original_index = request.name

    request_message = (
        f"**Заявка №{escape_markdown_v2(str(index + 1))} из {escape_markdown_v2(str(total_requests))}**\n"
        f"Имя: {escape_markdown_v2(request['Имя'])}\n"
        f"Номер: {escape_markdown_v2(request['Номер'])}\n"
        f"Выбор: {escape_markdown_v2(request['Выбор'])}\n"
        f"Дата: {escape_markdown_v2(request['Дата обращения'])}\n"
    )

    keyboard_buttons = []

    if index > 0:
        keyboard_buttons.append(InlineKeyboardButton(text="⬅️ Предыдущая", callback_data=f"show_inbox_{index - 1}"))
    
    keyboard_buttons.append(InlineKeyboardButton(text="✅ Обработано", callback_data=f"status_done_{original_index}"))

    if index < total_requests - 1:
        keyboard_buttons.append(InlineKeyboardButton(text="➡️ Следующая", callback_data=f"show_inbox_{index + 1}"))
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[keyboard_buttons])

    try:
        if message_id:
            await bot.edit_message_text(chat_id=chat_id, message_id=message_id, text=request_message, reply_markup=keyboard, parse_mode=ParseMode.MARKDOWN_V2)
        else:
            await bot.send_message(chat_id=chat_id, text=request_message, reply_markup=keyboard, parse_mode=ParseMode.MARKDOWN_V2)
    except (TelegramBadRequest, ValidationError) as e:
        print(f"Ошибка при редактировании/отправке сообщения: {e}")
        try:
            await bot.send_message(chat_id=chat_id, text=f"Произошла ошибка при обновлении\\. Подробности смотрите в консоли\\.", parse_mode=ParseMode.MARKDOWN_V2)
        except (TelegramBadRequest, ValidationError) as e2:
            print(f"Ошибка при отправке сообщения об ошибке: {e2}")


# --- ФУНКЦИИ ДЛЯ ПАНЕЛЕЙ ---
async def _send_db_backup(chat_id: int, bot: Bot):
    """Создает zip-архив базы данных и отправляет его."""
    try:
        if not os.path.exists(DB_FILE):
            await bot.send_message(chat_id=chat_id, text="Файл базы данных не найден.")
            return

        zip_filename = f"clients_backup_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.zip"
        with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(DB_FILE, os.path.basename(DB_FILE))

        try:
            await bot.send_document(chat_id=chat_id, document=FSInputFile(zip_filename), caption="✅ Бэкап базы данных успешно создан.")
        except (TelegramBadRequest, ValidationError) as e:
            print(f"Ошибка при отправке бэкапа: {e}")
            await bot.send_message(chat_id=chat_id, text=f"Произошла ошибка при отправке бэкапа: {escape_markdown_v2(str(e))}", parse_mode=ParseMode.MARKDOWN_V2)


        os.remove(zip_filename)
        print(f"Бэкап {zip_filename} отправлен и удален.")

    except Exception as e:
        print(f"Ошибка при создании или отправке бэкапа: {e}")
        try:
            await bot.send_message(chat_id=chat_id, text=f"Произошла ошибка при создании бэкапа: {escape_markdown_v2(str(e))}", parse_mode=ParseMode.MARKDOWN_V2)
        except (TelegramBadRequest, ValidationError) as e:
            print(f"Ошибка при отправке сообщения об ошибке: {e}")


async def _generate_stats_report(chat_id: int, message_id: int, bot: Bot, days: int = None):
    """Собирает и отправляет статистику по заявкам за выбранный период."""
    try:
        df = pd.read_excel(DB_FILE)

        if days is not None:
            time_filter = datetime.now() - timedelta(days=days)
            df['Дата обращения'] = pd.to_datetime(df['Дата обращения'])
            filtered_df = df[df['Дата обращения'] > time_filter]
        else:
            filtered_df = df.copy()

        if days == 365:
            period_text = "за последний год"
        elif days is not None:
            period_text = f"за последние {days} дней"
        else:
            period_text = "за все время"

        total_requests = len(filtered_df)
        unprocessed_requests = len(filtered_df[filtered_df['Статус'] == 'Необработана'])

        stats_message = (
            f"**Статистика по заявкам {period_text}**\n"
            f"Всего заявок: {escape_markdown_v2(total_requests)}\n"
            f"Необработанных: {escape_markdown_v2(unprocessed_requests)}\n"
        )
        markup = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ Назад", callback_data="open_stats_menu")]
        ])

        try:
            await bot.edit_message_text(
                text=stats_message,
                chat_id=chat_id,
                message_id=message_id,
                parse_mode=ParseMode.MARKDOWN_V2,
                reply_markup=markup
            )
        except (TelegramBadRequest, ValidationError) as e:
            print(f"Ошибка при редактировании статистики: {e}")
            await bot.send_message(chat_id=chat_id, text=f"Произошла ошибка при обновлении: {escape_markdown_v2(str(e))}", parse_mode=ParseMode.MARKDOWN_V2)
    except FileNotFoundError:
        try:
            await bot.edit_message_text(text="Файл базы данных не найден\\. Статистика недоступна\\.", chat_id=chat_id, message_id=message_id, parse_mode=ParseMode.MARKDOWN_V2)
        except (TelegramBadRequest, ValidationError) as e:
            print(f"Ошибка при редактировании сообщения 'файл не найден': {e}")
    except Exception as e:
        try:
            await bot.edit_message_text(text=f"Ошибка при сборе статистики: {escape_markdown_v2(str(e))}", chat_id=chat_id, message_id=message_id, parse_mode=ParseMode.MARKDOWN_V2)
        except (TelegramBadRequest, ValidationError) as e:
            print(f"Ошибка при редактировании сообщения 'ошибка сбора статистики': {e}")
        print(f"Ошибка при сборе статистики: {e}")

def create_dynamic_keyboard(items: list, callback_prefix: str) -> InlineKeyboardMarkup:
    """
    Создает Inline-клавиатуру, автоматически подстраивая количество кнопок
    в ряду под длину текста, чтобы избежать обрезки.
    """
    markup = InlineKeyboardMarkup(inline_keyboard=[])
    buttons = [InlineKeyboardButton(text=name, callback_data=f"{callback_prefix}{value}")
               for item_dict in items for name, value in item_dict.items()]

    current_row = []
    current_length = 0
    max_length_per_row = 30 # Примерное ограничение, можно настроить

    for button in buttons:
        button_length = len(button.text) + 5
        if current_length + button_length < max_length_per_row:
            current_row.append(button)
            current_length += button_length
        else:
            if current_row:
                markup.inline_keyboard.append(current_row)
            current_row = [button]
            current_length = button_length

    if current_row:
        markup.inline_keyboard.append(current_row)

    return markup


# --- ХЕНДЛЕРЫ ---
router = Router()


@router.message(Command('start'))
async def handle_start(message: Message):
    """Отправляет приветствие и кнопку для запроса контакта."""
    keyboard = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True, keyboard=[
        [KeyboardButton(text="Поделиться контактом", request_contact=True)]
    ])
    try:
        await message.answer(
            'Привет\\! 👋\n\nНаш бот поможет вам подобрать подходящий курс\\. Для начала, пожалуйста, нажмите кнопку 👇, чтобы мы могли с вами связаться\\.',
            reply_markup=keyboard,
            parse_mode=ParseMode.MARKDOWN_V2
        )
    except (TelegramBadRequest, ValidationError) as e:
        print(f"Ошибка при отправке сообщения 'start': {e}")
        await message.answer(
            f"Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}",
            parse_mode=ParseMode.MARKDOWN_V2
        )


@router.message(Command('admin'))
async def handle_admin_panel(message: Message):
    if str(message.from_user.id).strip() != str(config.admin_id).strip():
        try:
            await message.reply("У вас нет прав для доступа к этой команде.")
        except (TelegramBadRequest, ValidationError) as e:
            print(f"Ошибка при отправке сообщения 'admin': {e}")
        return

    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Показать заявки", callback_data="show_inbox")],
        [InlineKeyboardButton(text="Связаться с разработчиком", url=f"tg://user?id={config.dev_id}")]
    ])

    try:
        await message.answer(
            "**Административная панель**\n\nВыберите действие:",
            reply_markup=markup,
            parse_mode=ParseMode.MARKDOWN_V2
        )
    except (TelegramBadRequest, ValidationError) as e:
        print(f"Ошибка Pydantic при отправке панели администратора: {e}")
        await message.answer(
            f"Произошла внутренняя ошибка при загрузке панели\\: {escape_markdown_v2(str(e))}",
            parse_mode=ParseMode.MARKDOWN_V2
        )


@router.message(Command('dev'))
async def handle_dev_panel(message: Message):
    if str(message.from_user.id).strip() != str(config.dev_id).strip():
        try:
            await message.reply("У вас нет прав для доступа к этой команде.")
        except (TelegramBadRequest, ValidationError) as e:
            print(f"Ошибка при отправке сообщения 'dev': {e}")
        return

    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Очистить БД", callback_data="clear_db_confirm")],
        [
            InlineKeyboardButton(text="Показать статистику", callback_data="open_stats_menu"),
            InlineKeyboardButton(text="Сделать бэкап", callback_data="make_backup"),
        ]
    ])

    try:
        await message.answer(
            "**Панель разработчика**\n\nВыберите действие:",
            reply_markup=markup,
            parse_mode=ParseMode.MARKDOWN_V2
        )
    except (TelegramBadRequest, ValidationError) as e:
        print(f"Ошибка при отправке панели разработчика: {e}")
        await message.answer(
            f"Произошла внутренняя ошибка при загрузке панели\\: {escape_markdown_v2(str(e))}",
            parse_mode=ParseMode.MARKDOWN_V2
        )


@router.message(Command('owner'))
async def handle_owner_panel(message: Message):
    if str(message.from_user.id).strip() != str(config.owner_id).strip():
        try:
            await message.reply("У вас нет прав для доступа к этой команде.")
        except (TelegramBadRequest, ValidationError) as e:
            print(f"Ошибка при отправке сообщения 'owner': {e}")
        return

    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Показать статистику", callback_data="open_stats_menu")]
    ])

    try:
        await message.answer(
            "**Панель владельца**\n\nВыберите действие:",
            reply_markup=markup,
            parse_mode=ParseMode.MARKDOWN_V2
        )
    except (TelegramBadRequest, ValidationError) as e:
        print(f"Ошибка при отправке панели владельца: {e}")
        await message.answer(
            f"Произошла внутренняя ошибка при загрузке панели\\: {escape_markdown_v2(str(e))}",
            parse_mode=ParseMode.MARKDOWN_V2
        )


@router.message(Command('my_id'))
async def get_my_id(message: Message):
    try:
        await message.reply(f"Ваш ID: {message.chat.id}")
    except (TelegramBadRequest, ValidationError) as e:
        print(f"Ошибка при отправке ID: {e}")
        await message.answer(f"Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}", parse_mode=ParseMode.MARKDOWN_V2)


@router.message(F.contact)
async def handle_contact(message: Message):
    """Временно сохраняет контакт пользователя и запускает опрос."""
    if not message.contact:
        try:
            await message.answer("Пожалуйста, поделитесь контактом, нажав на кнопку.")
        except (TelegramBadRequest, ValidationError) as e:
            print(f"Ошибка при отправке сообщения 'contact' (нет контакта): {e}")
        return

    user_id = message.from_user.id
    phone_number = ''.join(filter(str.isdigit, message.contact.phone_number))
    first_name = message.contact.first_name

    user_data[user_id] = {
        'name': first_name,
        'phone': phone_number
    }
    print(f"Временные данные сохранены для user_id {user_id}: {user_data[user_id]}")

    try:
        await message.answer(
            "Спасибо, ваш контакт принят\\! ✅ Теперь, пожалуйста, ответьте на несколько вопросов\\.",
            reply_markup=ReplyKeyboardRemove(),
            parse_mode=ParseMode.MARKDOWN_V2
        )
    except (TelegramBadRequest, ValidationError) as e:
        print(f"Ошибка при отправке сообщения 'contact' (спасибо): {e}")
        await message.answer(
            f"Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}",
            parse_mode=ParseMode.MARKDOWN_V2
        )
        return

    markup = create_dynamic_keyboard(LANGS, callback_prefix='&s&l')

    try:
        await message.answer(
            'Выберите язык:',
            reply_markup=markup,
            parse_mode=ParseMode.MARKDOWN_V2
        )
    except (TelegramBadRequest, ValidationError) as e:
        print(f"Ошибка при отправке сообщения 'contact' (выбор языка): {e}")
        await message.answer(
            f"Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}",
            parse_mode=ParseMode.MARKDOWN_V2
        )


@router.callback_query(F.data == "open_stats_menu")
async def open_stats_menu(call: CallbackQuery):
    if str(call.from_user.id).strip() not in [str(config.owner_id).strip(), str(config.dev_id).strip()]:
        await call.answer("У вас нет прав.")
        return

    markup = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="За 24 часа", callback_data="stats_1d"),
            InlineKeyboardButton(text="За 7 дней", callback_data="stats_7d"),
        ],
        [
            InlineKeyboardButton(text="За 30 дней", callback_data="stats_30d"),
            InlineKeyboardButton(text="За год", callback_data="stats_1y"),
        ],
        [InlineKeyboardButton(text="За все время", callback_data="stats_all")]
    ])

    try:
        await call.message.edit_text(
            "Выберите период для статистики:",
            reply_markup=markup
        )
    except (TelegramBadRequest, ValidationError) as e:
        print(f"Ошибка при редактировании меню статистики: {e}")
        try:
            await call.message.answer(
                f"Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}",
                parse_mode=ParseMode.MARKDOWN_V2
            )
        except (TelegramBadRequest, ValidationError) as e2:
            print(f"Ошибка при отправке сообщения об ошибке: {e2}")
    await call.answer()


@router.callback_query(F.data.startswith('stats_'))
async def handle_stats_choice(call: CallbackQuery, bot: Bot):
    if str(call.from_user.id).strip() not in [str(config.owner_id).strip(), str(config.dev_id).strip()]:
        await call.answer("У вас нет прав.")
        return

    period = call.data.split('_')[1]

    days = None
    if period == '1d': days = 1
    elif period == '7d': days = 7
    elif period == '30d': days = 30
    elif period == '1y': days = 365

    await call.answer("Генерирую отчёт.")
    await _generate_stats_report(call.message.chat.id, call.message.message_id, bot, days)


@router.callback_query(F.data.regexp(r'^&s&l\d+$'))
async def handle_language_choice(call: CallbackQuery):
    """Обрабатывает выбор языка и предлагает выбрать формат (онлайн/очно)."""
    
    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=name, callback_data=f"{call.data}&d{value}") for delivery_dict in DELIVERY_METHODS for name, value in delivery_dict.items()],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data=call.data.rsplit('&', 1)[0])]
    ])

    try:
        await call.message.edit_text(
            'Выберите формат занятий:',
            reply_markup=markup,
            parse_mode=ParseMode.MARKDOWN_V2
        )
    except (TelegramBadRequest, ValidationError) as e:
        print(f"Ошибка при редактировании 'language_choice': {e}")
        await call.message.answer(
            f"Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}",
            parse_mode=ParseMode.MARKDOWN_V2
        )
    await call.answer()

@router.callback_query(F.data == '&s')
async def handle_back_to_langs(call: CallbackQuery):
    """
    Обрабатывает возврат к выбору языка с первого шага. Не запрашивает контакт повторно.
    """
    markup = create_dynamic_keyboard(LANGS, callback_prefix='&s&l')
    
    try:
        await call.message.edit_text(
            'Выберите язык:',
            reply_markup=markup,
            parse_mode=ParseMode.MARKDOWN_V2
        )
    except (TelegramBadRequest, ValidationError) as e:
        print(f"Ошибка при редактировании 'back_to_langs': {e}")
        await call.message.answer(
            f"Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}",
            parse_mode=ParseMode.MARKDOWN_V2
        )
    await call.answer()
    

@router.callback_query(F.data.regexp(r'.*&d\d+$'))
async def handle_delivery_choice(call: CallbackQuery):
    """Обрабатывает выбор формата и предлагает выбрать цель обучения."""
    markup = create_dynamic_keyboard(GOALS, callback_prefix=call.data + '&o')
    markup.inline_keyboard.append([InlineKeyboardButton(text="⬅️ Назад", callback_data=call.data.rsplit('&', 1)[0])])
    
    try:
        await call.message.edit_text(
            'Выберите цель обучения:',
            reply_markup=markup,
            parse_mode=ParseMode.MARKDOWN_V2
        )
    except (TelegramBadRequest, ValidationError) as e:
        print(f"Ошибка при редактировании 'delivery_choice': {e}")
        await call.message.answer(
            f"Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}",
            parse_mode=ParseMode.MARKDOWN_V2
        )
    await call.answer()


@router.callback_query(F.data.regexp(r'.*&o\d+$'))
async def handle_goal_choice(call: CallbackQuery):
    """Обрабатывает выбор цели и предлагает выбрать тип группы."""
    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=name, callback_data=f"{call.data}&g{value}")]
        for group_dict in GROUPS for name, value in group_dict.items()
    ])
    
    markup.inline_keyboard.append([InlineKeyboardButton(text="⬅️ Назад", callback_data=call.data.rsplit('&', 1)[0])])

    try:
        await call.message.edit_text(
            'Выберите формат занятий:',
            reply_markup=markup,
            parse_mode=ParseMode.MARKDOWN_V2
        )
    except (TelegramBadRequest, ValidationError) as e:
        print(f"Ошибка при редактировании 'goal_choice': {e}")
        await call.message.answer(
            f"Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}",
            parse_mode=ParseMode.MARKDOWN_V2
        )
    await call.answer()


@router.callback_query(F.data.regexp(r'.*&g\d+$'))
async def handle_group_choice(call: CallbackQuery):
    """Обрабатывает выбор группы и предлагает выбрать возрастную группу."""
    markup = create_dynamic_keyboard(AGES, callback_prefix=call.data + '&a')
    markup.inline_keyboard.append([InlineKeyboardButton(text="⬅️ Назад", callback_data=call.data.rsplit('&', 1)[0])])

    try:
        await call.message.edit_text(
            'Выберите возрастную группу:',
            reply_markup=markup,
            parse_mode=ParseMode.MARKDOWN_V2
        )
    except (TelegramBadRequest, ValidationError) as e:
        print(f"Ошибка при редактировании 'group_choice': {e}")
        await call.message.answer(
            f"Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}",
            parse_mode=ParseMode.MARKDOWN_V2
        )
    await call.answer()


@router.callback_query(F.data.regexp(r'.*&a\d+$'))
async def handle_age_choice(call: CallbackQuery):
    """Обрабатывает выбор возраста и предлагает выбрать уровень."""
    markup = create_dynamic_keyboard(LEVELS, callback_prefix=call.data + '&e')
    markup.inline_keyboard.append([InlineKeyboardButton(text="⬅️ Назад", callback_data=call.data.rsplit('&', 1)[0])])
    
    try:
        await call.message.edit_text(
            'Выберите ваш уровень:',
            reply_markup=markup,
            parse_mode=ParseMode.MARKDOWN_V2
        )
    except (TelegramBadRequest, ValidationError) as e:
        print(f"Ошибка при редактировании 'age_choice': {e}")
        await call.message.answer(
            f"Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}",
            parse_mode=ParseMode.MARKDOWN_V2
        )
    await call.answer()


@router.callback_query(F.data.regexp(r'.*&e\d+$'))
async def handle_final_choice(call: CallbackQuery, bot: Bot):
    user_id = call.from_user.id
    if user_id not in user_data:
        try:
            await call.message.edit_text("Произошла ошибка\\. Пожалуйста, начните сначала с команды /start\\.", parse_mode=ParseMode.MARKDOWN_V2)
        except (TelegramBadRequest, ValidationError) as e:
            print(f"Ошибка при редактировании 'final_choice': {e}")
            await call.message.answer(
                f"Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}",
                parse_mode=ParseMode.MARKDOWN_V2
            )
        await call.answer("Произошла ошибка.")
        return

    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    decoded_choice = decode_selection(call.data)
    name = user_data[user_id]['name']
    phone = user_data[user_id]['phone']
    new_record = {
        'Имя': name,
        'Номер': phone,
        'Дата обращения': timestamp,
        'Выбор': decoded_choice,
        'Статус': 'Необработана'
    }

    try:
        is_file_new = not os.path.exists(DB_FILE)
        df = pd.read_excel(DB_FILE, dtype=str)
    except FileNotFoundError:
        df = pd.DataFrame(columns=['Имя', 'Номер', 'Дата обращения', 'Выбор', 'Статус'])

    is_duplicate = df[(df['Номер'] == phone) & (df['Выбор'] == decoded_choice)].shape[0] > 0

    if is_duplicate:
        try:
            await call.message.edit_text("Мы уже приняли вашу заявку\\. Наш менеджер скоро свяжется с вами\\. ✅", parse_mode=ParseMode.MARKDOWN_V2)
        except (TelegramBadRequest, ValidationError) as e:
            print(f"Ошибка при редактировании 'duplicate': {e}")
            await call.message.answer(
                f"Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}",
                parse_mode=ParseMode.MARKDOWN_V2
            )
    else:
        df_new = pd.DataFrame([new_record])
        df = pd.concat([df, df_new], ignore_index=True)
        new_record_index = df.tail(1).index[0]

        with pd.ExcelWriter(DB_FILE, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            for column in worksheet.columns:
                max_length = max((len(str(cell.value)) for cell in column if cell.value), default=10)
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        admin_requests_cache.clear() # ОЧИЩАЕМ КЭШ ПОСЛЕ ЗАПИСИ
        print(f"Новая запись успешно сохранена в {DB_FILE} с автонастройкой ширины.")

         # Проверка перед отправкой на отправку в рабочее время
        if is_within_working_hours():
            admin_message = (
                f"🔔 **НОВАЯ ЗАЯВКА**\n"
                f"Имя: {escape_markdown_v2(name)}\n"
                f"Номер: {escape_markdown_v2(phone)}\n"
                f"Выбор: {escape_markdown_v2(decoded_choice)}\n"
                f"Дата и время обращения: {escape_markdown_v2(timestamp)}"
            )
            try:
                await bot.send_message(chat_id=config.admin_id, text=admin_message, parse_mode=ParseMode.MARKDOWN_V2)
            except (TelegramBadRequest, ValidationError) as e:
                print(f"Ошибка при отправке сообщения администратору: {e}")
                await bot.send_message(chat_id=config.admin_id, text=f"❌ Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}", parse_mode=ParseMode.MARKDOWN_V2)

            # Теперь этот вызов только в рабочее время
            asyncio.create_task(send_reminder_to_admin(bot, config.admin_id, name, phone, new_record_index, 1))
        else:
            # Сохранение заявки, для отправки в рабочее время
            queued_request = {
                'name': name,
                'phone': phone,
                'choice': decoded_choice,
                'timestamp': timestamp
            }
            add_to_queue(queued_request)
            print(f"Заявка добавлена в очередь, так как время нерабочее.")

        final_message = (
            f"Спасибо\\! 🎉\n\n"
            f"Ваша заявка принята\\. *Скоро* наш менеджер свяжется с вами, чтобы обсудить детали по курсу\\.\n\n"
            f"👉 *Ваш выбор:*\n"
            f"{escape_markdown_v2(decoded_choice)}\n"
            f"Ожидайте звонка\\! 📞"
        )
        # Удаляем предыдущее сообщение с кнопками
        try:
            await bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)
        except TelegramBadRequest as e:
            print(f"Ошибка при удалении сообщения: {e}")

        # Отправляем сообщение с текстом и геолокацией
        await bot.send_message(
            chat_id=call.message.chat.id, 
            text=final_message, 
            parse_mode=ParseMode.MARKDOWN_V2
        )
        
        # Отправляем геолокацию
        await bot.send_venue(
            chat_id=call.message.chat.id, 
            latitude=config.venue_lat,
            longitude=config.venue_lon,
            title=config.venue_title,
            address=config.venue_address
        )

    del user_data[user_id]
    await call.answer()

async def process_request_queue(bot: Bot):
    """Обрабатывает очередь запросов и отправляет сводку администратору."""
    # This task will run daily, you can define the exact timing in main()
    now = datetime.now(MOSCOW_TZ)
    
    # Check if it's the right time to send the summary (e.g., daily at 10:05 AM)
    if now.hour == config.ADMIN_WORKING_HOURS['mon_fri_start'] and now.minute == 5:
        if not os.path.exists(config.REQUEST_QUEUE_FILE):
            print("Очередь заявок пуста.")
            return

        try:
            with open(config.REQUEST_QUEUE_FILE, 'r', encoding='utf-8') as f:
                queue = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            print("Файл очереди не найден или поврежден.")
            return

        if not queue:
            print("Очередь заявок пуста.")
            return

        summary_message = f"**Сводка по заявкам за нерабочее время**\n\n"
        for request in queue:
            summary_message += (
                f"**Имя:** {escape_markdown_v2(request['name'])}\n"
                f"**Номер:** {escape_markdown_v2(request['phone'])}\n"
                f"**Выбор:** {escape_markdown_v2(request['choice'])}\n"
                f"**Дата и время:** {escape_markdown_v2(request['timestamp'])}\n"
                f"\\-\\-\\-\n"
            )

        try:
            await bot.send_message(
                chat_id=config.admin_id,
                text=summary_message,
                parse_mode=ParseMode.MARKDOWN_V2
            )
            print("Сводка по заявкам успешно отправлена.")
            clear_queue() # Очистка очереди после отправки
        except (TelegramBadRequest, ValidationError) as e:
            print(f"Ошибка при отправке сводки: {e}")

@router.callback_query(F.data.in_(["show_inbox", "clear_db_confirm", "make_backup"]))
async def handle_admin_buttons(call: CallbackQuery, bot: Bot):
    if str(call.from_user.id).strip() not in [str(config.admin_id).strip(), str(config.dev_id).strip(), str(config.owner_id).strip()]:
        await call.answer("У вас нет прав.")
        return

    if call.data == "show_inbox":
        if str(call.from_user.id).strip() == str(config.admin_id).strip():
            await call.answer("Загрузка заявок.")
            await _send_inbox_messages(call.message.chat.id, bot, message_id=call.message.message_id)
        else:
            await call.answer("У вас нет прав для просмотра заявок.")
    elif call.data == "clear_db_confirm":
        if str(call.from_user.id).strip() == str(config.dev_id).strip():
            markup = InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(text="Да, очистить", callback_data="confirm_clear_yes"),
                    InlineKeyboardButton(text="Отмена", callback_data="confirm_clear_no")
                ]
            ])
            try:
                await call.message.edit_text(
                    " **ВНИМАНИЕ** \n\nВы уверены, что хотите полностью очистить базу данных заявок\\? Это действие необратимо\\!",
                    reply_markup=markup,
                    parse_mode=ParseMode.MARKDOWN_V2
                )
            except (TelegramBadRequest, ValidationError) as e:
                print(f"Ошибка при редактировании 'clear_db_confirm': {e}")
                await call.message.answer(
                    f"Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}",
                    parse_mode=ParseMode.MARKDOWN_V2
                )
        else:
            await call.answer("У вас нет прав для очистки БД.")
    elif call.data == "make_backup":
        if str(call.from_user.id).strip() == str(config.dev_id).strip():
            await call.answer("Создаю бэкап.")
            await _send_db_backup(call.message.chat.id, bot)
        else:
            await call.answer("У вас нет прав для создания бэкапа.")

    await call.answer()


@router.callback_query(F.data.startswith('show_inbox_'))
async def handle_show_inbox(call: CallbackQuery, bot: Bot):
    if str(call.from_user.id).strip() != str(config.admin_id).strip():
        await call.answer("У вас нет прав для просмотра заявок.")
        return
    
    index = int(call.data.split('_')[2])
    
    await call.answer("Загрузка.")
    await _send_inbox_messages(call.message.chat.id, bot, index, call.message.message_id)


@router.callback_query(F.data.startswith('confirm_clear_'))
async def handle_clear_confirmation(call: CallbackQuery):
    if str(call.from_user.id).strip() != str(config.dev_id).strip():
        await call.answer("У вас нет прав.")
        return

    confirmation = call.data.split('_')[-1]

    if confirmation == 'yes':
        try:
            df_empty = pd.DataFrame(columns=['Имя', 'Номер', 'Дата обращения', 'Выбор', 'Статус'])
            with pd.ExcelWriter(DB_FILE, engine='openpyxl') as writer:
                df_empty.to_excel(writer, index=False)

            try:
                await call.message.edit_text(
                    "База данных заявок успешно очищена\\.",
                    parse_mode=ParseMode.MARKDOWN_V2,
                    reply_markup=None
                )
            except (TelegramBadRequest, ValidationError) as e:
                print(f"Ошибка при редактировании 'clear_confirmation_yes': {e}")
                await call.message.answer(
                    f"Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}",
                    parse_mode=ParseMode.MARKDOWN_V2
                )
            print("База данных успешно очищена.")
        except Exception as e:
            try:
                await call.message.edit_text(f"Произошла ошибка при очистке: {escape_markdown_v2(str(e))}", parse_mode=ParseMode.MARKDOWN_V2)
            except (TelegramBadRequest, ValidationError) as e2:
                print(f"Ошибка при отправке сообщения об ошибке: {e2}")
            print(f"Ошибка при очистке базы данных: {e}")

    elif confirmation == 'no':
        try:
            await call.message.edit_text("Очистка базы данных отменена\\.", parse_mode=ParseMode.MARKDOWN_V2, reply_markup=None)
        except (TelegramBadRequest, ValidationError) as e:
            print(f"Ошибка при редактировании 'clear_confirmation_no': {e}")
            await call.message.answer(
                f"Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}",
                parse_mode=ParseMode.MARKDOWN_V2
            )
        print("Очистка базы данных отменена.")

    await call.answer()


@router.callback_query(F.data == "start_over")
async def handle_start_over(call: CallbackQuery):
    """Возвращает пользователя в начало опроса."""
    try:
        await handle_start(call.message)
    except Exception as e:
        print(f"Ошибка при запуске 'start_over': {e}")
    finally:
        await call.answer()


@router.callback_query(F.data.startswith('status_'))
async def handle_status_change(call: CallbackQuery, bot: Bot):
    if str(call.from_user.id).strip() != str(config.admin_id).strip():
        await call.answer("У вас нет прав.")
        return

    try:
        _, status, index_str = call.data.split('_')
        original_index = int(index_str)
        
        # Update status in Excel file
        df = pd.read_excel(DB_FILE)
        df.loc[original_index, 'Статус'] = 'Обработано'
        df.to_excel(DB_FILE, index=False)
        
        # Clear cache to force refresh
        admin_requests_cache.clear()

        await call.answer("Статус заявки обновлен.")

        await _send_inbox_messages(call.message.chat.id, bot, 0, call.message.message_id)

    except Exception as e:
        await call.answer("Произошла ошибка при обновлении.")
        print(f"Ошибка при обновлении статуса: {e}")


async def send_reminder_to_admin(bot: Bot, chat_id: int, name: str, phone: str, index: int, delay_minutes: int):
    """Отправляет уведомление администратору о заявке, если ее статус 'Необработана'."""
    delay_minutes = 5
    delay_seconds = delay_minutes * 60
    await asyncio.sleep(delay_seconds)

    try:
        df = pd.read_excel(DB_FILE)
        if index in df.index:
            current_status = df.loc[index, 'Статус']
            if current_status == 'Необработана':
                reminder_message = (
                    f"⚠️ **НАПОМИНАНИЕ**\n"
                    f"Заявка от клиента *{escape_markdown_v2(name)}* всё ещё не обработана\\.\n"
                    f"Номер: {escape_markdown_v2(phone)}"
                )
                try:
                    await bot.send_message(chat_id=chat_id, text=reminder_message, parse_mode=ParseMode.MARKDOWN_V2)
                except (TelegramBadRequest, ValidationError) as e:
                    print(f"Ошибка при отправке напоминания: {e}")
                    await bot.send_message(chat_id=chat_id, text=f"❌ Произошла внутренняя ошибка: {escape_markdown_v2(str(e))}", parse_mode=ParseMode.MARKDOWN_V2)
                print(f"Отправлено напоминание администратору о заявке {index}.")
            else:
                print(f"Заявка {index} уже обработана (Статус: {current_status}). Напоминание не отправлено.")
        else:
            print(f"Ошибка: Индекс {index} не найден в файле Excel. Напоминание не отправлено.")
    except Exception as e:
        print(f"Ошибка при проверке статуса и отправке напоминания: {e}")

# --- ЗАПУСК БОТА ---
# In main.py, modify the main function
async def main():
    bot = Bot(token=config.token)
    dp = Dispatcher()
    dp.include_router(router)

    print('Бот запущен...')

    # Start the queue processing task in the background
    asyncio.create_task(periodic_queue_check(bot))
    
    await dp.start_polling(bot)


async def periodic_queue_check(bot: Bot):
    """
    Планирует задачу обработки очереди и отправляет ежедневные приветствия.
    Утреннее приветствие отправляется в начале рабочего дня + 5 минут, а вечернее - в конце рабочего дня + 5 минут.
    """
    last_morning_greeting_sent = None
    last_evening_greeting_sent = None
    
    while True:
        now = datetime.now(MOSCOW_TZ)
        today = now.date()
        
        # Получаем часы начала и конца рабочего дня из конфига
        start_hour = config.ADMIN_WORKING_HOURS['mon_fri_start'] if 0 <= now.weekday() <= 4 else config.ADMIN_WORKING_HOURS['sat_sun_start']
        end_hour = config.ADMIN_WORKING_HOURS['mon_fri_end'] if 0 <= now.weekday() <= 4 else config.ADMIN_WORKING_HOURS['sat_sun_end']

        # Утреннее приветствие и обработка очереди
        # Отправляем в начале рабочего дня + 5 минут
        if now.hour == start_hour and now.minute == 5 and last_morning_greeting_sent != today:
            greeting_message = "Доброе утро! 🌅 Сегодня начинается рабочий день, и я готов принимать заявки."
            try:
                await bot.send_message(chat_id=config.admin_id, text=greeting_message)
                await process_request_queue(bot)  # Обрабатываем все накопленные заявки
                last_morning_greeting_sent = today
            except Exception as e:
                print(f"Ошибка при отправке утреннего приветствия или обработке очереди: {e}")

        # --- Вечернее сообщение ---
        # Отправляем в конце рабочего дня + 5 минут
        if now.hour == end_hour and now.minute == 5 and last_evening_greeting_sent != today:
            farewell_message = "Рабочий день окончен! 🌙 Спасибо за вашу работу, хорошего вечера и до завтра!"
            try:
                await bot.send_message(chat_id=config.admin_id, text=farewell_message)
                last_evening_greeting_sent = today
            except Exception as e:
                print(f"Ошибка при отправке вечернего сообщения: {e}")

        # Сон на одну минуту, чтобы избежать повторной отправки
        await asyncio.sleep(60)


if __name__ == '__main__':
    if not os.path.exists('db'):
        os.makedirs('db')
    asyncio.run(main())
    
